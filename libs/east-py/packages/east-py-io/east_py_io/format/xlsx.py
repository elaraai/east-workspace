#
# Copyright (c) 2025 Elara AI Pty Ltd
# Licensed under the Business Source License 1.1. See LICENSE.md for details.
#
"""XLSX platform functions for East.

The ``*_impl`` functions are plain Python callables taking and returning East
values - import them directly from a project's own ``@platform_function`` to
reuse the implementations without an IR round-trip. Requires the ``xlsx``
extra (``openpyxl``).
"""

import importlib.util
import io
from datetime import UTC, datetime
from typing import Any

from east.runtime.platform import platform_function, platform_functions
from east.types.types import BlobType
from east.types.values import EastArray, EastBlob, EastStruct, EastVariant, east_null

_HAS_XLSX_SUPPORT = importlib.util.find_spec("openpyxl") is not None


def _check_xlsx_support() -> None:
    """Raise if openpyxl is not installed."""
    if not _HAS_XLSX_SUPPORT:
        raise NotImplementedError(
            "XLSX support requires the 'xlsx' extra. "
            "Add east-py-io[xlsx] to your pyproject.toml dependencies."
        )

from .types import (
    LiteralValueType,
    XlsxInfoType,
    XlsxReadOptionsType,
    XlsxRowType,
    XlsxSheetInfoType,
    XlsxSheetType,
    XlsxWriteOptionsType,
)


def convert_cell_to_east(value: Any, data_type: str | None = None) -> EastVariant:
    """Convert an Excel cell value to an East ``LiteralValueType`` variant.

    Excel stores all numbers as floats internally, so integer-looking cells
    are returned as ``Float`` to match TypeScript behaviour.

    Args:
        value: The raw cell value from openpyxl.
        data_type: The cell's ``data_type`` attribute (``'s'``,
            ``'inlineStr'``, ``'n'``, etc.) used to detect empty strings
            stored as ``None``.
    """
    if value is None:
        if data_type in ("s", "inlineStr"):
            return EastVariant("String", "")
        return EastVariant("Null", east_null)
    elif isinstance(value, bool):
        return EastVariant("Boolean", value)
    elif isinstance(value, int):
        return EastVariant("Float", float(value))
    elif isinstance(value, float):
        return EastVariant("Float", value)
    elif isinstance(value, datetime):
        value = value.replace(tzinfo=UTC) if value.tzinfo is None else value.astimezone(UTC)
        ms = (value.microsecond // 1000) * 1000
        value = value.replace(microsecond=ms)
        return EastVariant("DateTime", value)
    elif isinstance(value, str):
        return EastVariant("String", value)
    else:
        return EastVariant("String", str(value))


def convert_east_to_cell(value: EastVariant) -> Any:
    """Convert an East ``LiteralValueType`` variant to an Excel cell value.

    Args:
        value: ``EastVariant`` with tag matching a ``LiteralValueType``
            case (``Null``, ``Boolean``, ``Integer``, ``Float``,
            ``String``, ``DateTime``, ``Blob``).
    """
    tag = value.type
    val = value.value

    if tag == "Null":
        return None
    elif tag == "Boolean":
        return val
    elif tag == "Integer":
        return int(val) if val is not None else 0
    elif tag == "DateTime":
        if val is not None and hasattr(val, "tzinfo") and val.tzinfo is not None:
            return val.replace(tzinfo=None)
        return val
    elif tag == "Blob":
        return val.hex() if hasattr(val, "hex") else str(val)
    else:
        return val


@platform_function(
    name="xlsx_read",
    inputs=[BlobType, XlsxReadOptionsType],
    output=XlsxSheetType,
)
def xlsx_read_impl(blob: EastBlob, options: EastStruct) -> EastArray:
    """Read a single sheet from an XLSX file into a 2-D array of cells.

    Args:
        blob: ``Blob`` (``EastBlob``) - the raw XLSX file bytes.
        options: ``XlsxReadOptionsType`` (``EastStruct``) with fields:

            - ``sheetName`` (``Option<String>``): sheet to read; defaults
              to the workbook's active sheet when absent.

    Returns:
        ``Array<Array<LiteralValueType>>`` (``EastArray``) - rows of
        cell values. Each cell is a ``LiteralValueType`` variant.
        Empty cells are ``Null`` unless the cell was stored as an empty
        string (``inlineStr``/``s`` data type), in which case
        ``String("")`` is returned.

    Raises:
        NotImplementedError: the ``xlsx`` extra (``openpyxl``) is not
            installed.
        RuntimeError: the specified ``sheetName`` does not exist in the
            workbook, or the file is not a valid XLSX.
    """
    _check_xlsx_support()
    from openpyxl import load_workbook

    try:
        sheet_name_opt = options["sheetName"]
        sheet_name = sheet_name_opt.value if sheet_name_opt.type == "some" else None

        wb = load_workbook(filename=io.BytesIO(bytes(blob)), read_only=False, data_only=True)

        try:
            ws = wb[sheet_name] if sheet_name else wb.active
        except KeyError:
            wb.close()
            raise Exception(f'Sheet "{sheet_name}" not found in workbook') from None

        if ws is None:
            wb.close()
            return EastArray(XlsxRowType, [])

        result: EastArray = EastArray(XlsxRowType, [])
        for row in ws.iter_rows():
            row_data: EastArray = EastArray(
                LiteralValueType, [convert_cell_to_east(cell.value, cell.data_type) for cell in row]
            )
            result.push_last(row_data)

        wb.close()
        return result
    except Exception as e:
        if "not found in workbook" in str(e):
            raise
        raise Exception(f"XLSX read failed: {e}") from e


@platform_function(
    name="xlsx_write",
    inputs=[XlsxSheetType, XlsxWriteOptionsType],
    output=BlobType,
)
def xlsx_write_impl(data: EastArray, options: EastStruct) -> EastBlob:
    """Write a 2-D array of cells to an XLSX file.

    Args:
        data: ``Array<Array<LiteralValueType>>`` (``EastArray``) - rows
            of ``LiteralValueType`` cell values to write.
        options: ``XlsxWriteOptionsType`` (``EastStruct``) with fields:

            - ``sheetName`` (``Option<String>``): name for the sheet;
              defaults to ``"Sheet1"`` when absent.

    Returns:
        ``Blob`` (``EastBlob``) - the XLSX file bytes.

    Raises:
        NotImplementedError: the ``xlsx`` extra (``openpyxl``) is not
            installed.
        RuntimeError: serialization fails.
    """
    _check_xlsx_support()
    from openpyxl import Workbook

    try:
        sheet_name_opt = options["sheetName"]
        sheet_name = sheet_name_opt.value if sheet_name_opt.type == "some" else "Sheet1"

        wb = Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet()
        ws.title = sheet_name

        for row_idx, row in enumerate(data, start=1):
            for col_idx, cell in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=convert_east_to_cell(cell))

        output = io.BytesIO()
        wb.save(output)
        wb.close()

        return EastBlob(output.getvalue())
    except Exception as e:
        raise Exception(f"XLSX write failed: {e}") from e


@platform_function(
    name="xlsx_info",
    inputs=[BlobType],
    output=XlsxInfoType,
)
def xlsx_info_impl(blob: EastBlob) -> EastStruct:
    """Return metadata about all sheets in an XLSX workbook.

    Args:
        blob: ``Blob`` (``EastBlob``) - the raw XLSX file bytes.

    Returns:
        ``XlsxInfoType`` (``EastStruct``): ``sheets``
        (``Array<XlsxSheetInfoType>``) - one entry per sheet with
        ``name`` (``String``), ``rowCount`` (``Integer``), and
        ``columnCount`` (``Integer``).

    Raises:
        NotImplementedError: the ``xlsx`` extra (``openpyxl``) is not
            installed.
        RuntimeError: the file is not a valid XLSX.
    """
    _check_xlsx_support()
    from openpyxl import load_workbook

    try:
        wb = load_workbook(filename=io.BytesIO(bytes(blob)), read_only=True)

        sheets: EastArray = EastArray(XlsxSheetInfoType, [])
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheets.push_last(
                EastStruct(
                    {
                        "name": sheet_name,
                        "rowCount": ws.max_row or 0,
                        "columnCount": ws.max_column or 0,
                    }
                )
            )

        wb.close()
        return EastStruct({"sheets": sheets})
    except Exception as e:
        raise Exception(f"XLSX info failed: {e}") from e


# Collected from the @platform_function decorations above.
xlsx_impl = platform_functions(__name__)

__all__ = [
    "xlsx_impl",
    "xlsx_read_impl",
    "xlsx_write_impl",
    "xlsx_info_impl",
]
